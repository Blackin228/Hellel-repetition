import csv
import openpyxl

data_list = [['', 'Id', 'Name', 'Tel']]

with open('dz_17.csv') as file:
    reader = csv.reader(file)
    for index, row in enumerate(reader):
        if index == 0:
            continue
        else:
            data_list.append([f'Person {index}', row[0], row[1], row[3]])

wb = openpyxl.Workbook()
ws = wb.create_sheet('First sheet')
wb.remove(wb['Sheet'])

for index_column, value_row in enumerate(data_list):
    for index_row, value in enumerate(value_row):
        ws.cell(row=index_row+1, column=index_column+1, value=value)

wb.save('dz_18.xlsx')
